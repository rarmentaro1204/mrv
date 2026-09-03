#!/usr/bin/env python3
"""
Converte data/db_prospect_master.csv in un file Excel (.xlsx) pronto per
azioni commerciali (import CRM, mail merge, filtri per provincia/priorità),
con lo stesso stile grafico usato negli altri database prospect MRV
(DB_Prospect_CRM_Veneto.xlsx, prospect_lombardia_crm.xlsx):

- Foglio "Database Prospect": intestazione blu scuro/bianco, colonne
  larghe a misura, riga 1 congelata, autofiltro, bordi sottili,
  colonna PRIORITA colorata (ALTA verde / MEDIA giallo / BASSA rosso).
- Colonna REGIONE aggiunta (derivata dalla provincia) per poter
  segmentare le azioni commerciali anche per regione, non solo provincia.
- Foglio "Riepilogo": conteggio aziende per provincia e per priorità.

Uso:
    python3 scripts/export_xlsx.py
    python3 scripts/export_xlsx.py --csv data/db_prospect_master.csv --out data/db_prospect_master.xlsx
"""
import argparse
import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Provincia -> Regione, stesso raggruppamento geografico di data/rotation_state.json
PROVINCIA_REGIONE = {
    "Aosta": "Valle d'Aosta",
    "Torino": "Piemonte", "Vercelli": "Piemonte", "Novara": "Piemonte",
    "Cuneo": "Piemonte", "Asti": "Piemonte", "Alessandria": "Piemonte",
    "Biella": "Piemonte", "Verbano-Cusio-Ossola": "Piemonte",
    "Genova": "Liguria", "Imperia": "Liguria", "La Spezia": "Liguria", "Savona": "Liguria",
    "Varese": "Lombardia", "Como": "Lombardia", "Sondrio": "Lombardia", "Milano": "Lombardia",
    "Bergamo": "Lombardia", "Brescia": "Lombardia", "Pavia": "Lombardia", "Cremona": "Lombardia",
    "Mantova": "Lombardia", "Lecco": "Lombardia", "Lodi": "Lombardia",
    "Monza e della Brianza": "Lombardia",
    "Bolzano": "Trentino-Alto Adige", "Trento": "Trentino-Alto Adige",
    "Verona": "Veneto", "Vicenza": "Veneto", "Belluno": "Veneto", "Treviso": "Veneto",
    "Venezia": "Veneto", "Padova": "Veneto", "Rovigo": "Veneto",
    "Udine": "Friuli-Venezia Giulia", "Gorizia": "Friuli-Venezia Giulia",
    "Trieste": "Friuli-Venezia Giulia", "Pordenone": "Friuli-Venezia Giulia",
    "Piacenza": "Emilia-Romagna", "Parma": "Emilia-Romagna", "Reggio Emilia": "Emilia-Romagna",
    "Modena": "Emilia-Romagna", "Bologna": "Emilia-Romagna", "Ferrara": "Emilia-Romagna",
    "Ravenna": "Emilia-Romagna", "Forli-Cesena": "Emilia-Romagna", "Rimini": "Emilia-Romagna",
    "Massa-Carrara": "Toscana", "Lucca": "Toscana", "Pistoia": "Toscana", "Prato": "Toscana",
    "Firenze": "Toscana", "Livorno": "Toscana", "Pisa": "Toscana", "Arezzo": "Toscana",
    "Siena": "Toscana", "Grosseto": "Toscana",
    "Perugia": "Umbria", "Terni": "Umbria",
    "Pesaro e Urbino": "Marche", "Ancona": "Marche", "Macerata": "Marche",
    "Fermo": "Marche", "Ascoli Piceno": "Marche",
    "Viterbo": "Lazio", "Rieti": "Lazio", "Roma": "Lazio", "Latina": "Lazio", "Frosinone": "Lazio",
    "Teramo": "Abruzzo", "Pescara": "Abruzzo", "Chieti": "Abruzzo", "L'Aquila": "Abruzzo",
    "Campobasso": "Molise", "Isernia": "Molise",
    "Caserta": "Campania", "Napoli": "Campania", "Avellino": "Campania",
    "Benevento": "Campania", "Salerno": "Campania",
    "Foggia": "Puglia", "Barletta-Andria-Trani": "Puglia", "Bari": "Puglia",
    "Taranto": "Puglia", "Brindisi": "Puglia", "Lecce": "Puglia",
    "Potenza": "Basilicata", "Matera": "Basilicata",
    "Cosenza": "Calabria", "Crotone": "Calabria", "Catanzaro": "Calabria",
    "Vibo Valentia": "Calabria", "Reggio Calabria": "Calabria",
    "Messina": "Sicilia", "Palermo": "Sicilia", "Trapani": "Sicilia", "Agrigento": "Sicilia",
    "Caltanissetta": "Sicilia", "Enna": "Sicilia", "Catania": "Sicilia",
    "Siracusa": "Sicilia", "Ragusa": "Sicilia",
    "Sassari": "Sardegna", "Nuoro": "Sardegna", "Oristano": "Sardegna",
    "Sud Sardegna": "Sardegna", "Cagliari": "Sardegna",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9") for _ in range(4)))

PRIORITA_FILL = {
    "ALTA": PatternFill("solid", fgColor="C6EFCE"),
    "MEDIA": PatternFill("solid", fgColor="FFEB9C"),
    "BASSA": PatternFill("solid", fgColor="FFC7CE"),
}

# Larghezza colonna per header conosciuti (fallback 20 per header non mappati)
COLUMN_WIDTHS = {
    "SETTORE": 22, "SOTTOCATEGORIA": 26, "AZIENDA": 28, "REGIONE": 14,
    "PROVINCIA": 14, "SITO": 26, "TELEFONO": 16, "MAIL_GENERICA": 26,
    "NOME_DECISORE": 20, "RUOLO": 22, "LINKEDIN_DECISORE": 26,
    "SEGNALI_CRM": 34, "PRIORITA": 10, "FONTE_VERIFICA": 34, "DATA_INSERIMENTO": 14,
}


def load_rows(csv_path: Path):
    with csv_path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames), [row for row in reader]


def build_workbook(header, rows):
    # Inserisce REGIONE subito dopo AZIENDA, come negli altri DB prospect MRV
    out_header = list(header)
    if "REGIONE" not in out_header and "AZIENDA" in out_header:
        out_header.insert(out_header.index("AZIENDA") + 1, "REGIONE")

    wb = Workbook()
    ws = wb.active
    ws.title = "Database Prospect"

    ws.append(out_header)
    for c in range(1, len(out_header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    priorita_col = out_header.index("PRIORITA") + 1 if "PRIORITA" in out_header else None
    provincia_counter = Counter()
    priorita_counter = Counter()

    for row in rows:
        provincia = (row.get("PROVINCIA") or "").strip()
        row["REGIONE"] = PROVINCIA_REGIONE.get(provincia, "n.d.")
        values = [row.get(h, "") for h in out_header]
        ws.append(values)
        r = ws.max_row
        for c in range(1, len(out_header) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        if priorita_col:
            priorita = (row.get("PRIORITA") or "").strip().upper()
            fill = PRIORITA_FILL.get(priorita)
            if fill:
                ws.cell(row=r, column=priorita_col).fill = fill
        provincia_counter[provincia or "n.d."] += 1
        priorita_counter[(row.get("PRIORITA") or "n.d.").strip().upper() or "n.d."] += 1

    for idx, h in enumerate(out_header, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(h, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Foglio di riepilogo
    rs = wb.create_sheet("Riepilogo")
    rs.append([f"TOTALE AZIENDE: {len(rows)}"])
    rs.append([])
    rs.append(["Per Provincia"])
    rs.append(["Categoria", "Conteggio"])
    for provincia, count in provincia_counter.most_common():
        rs.append([provincia, count])
    rs.append([])
    rs.append(["Per Priorità"])
    rs.append(["Categoria", "Conteggio"])
    for priorita in ("ALTA", "MEDIA", "BASSA"):
        if priorita in priorita_counter:
            rs.append([priorita, priorita_counter[priorita]])
    for priorita, count in priorita_counter.items():
        if priorita not in ("ALTA", "MEDIA", "BASSA"):
            rs.append([priorita, count])
    rs.column_dimensions["A"].width = 26
    rs.column_dimensions["B"].width = 14
    for row in rs.iter_rows():
        for cell in row:
            cell.font = DATA_FONT
    for cell in ("A1", "A3", "A4", "B4"):
        rs[cell].font = Font(name="Arial", size=10, bold=True)
    header_row_idx = 5 + len(provincia_counter)  # riga "Per Priorità"
    rs[f"A{header_row_idx}"].font = Font(name="Arial", size=10, bold=True)

    return wb


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", default="data/db_prospect_master.csv", type=Path)
    parser.add_argument("--out", default="data/db_prospect_master.xlsx", type=Path)
    args = parser.parse_args()

    header, rows = load_rows(args.csv)
    wb = build_workbook(header, rows)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Scritte {len(rows)} aziende in {args.out}")


if __name__ == "__main__":
    main()
